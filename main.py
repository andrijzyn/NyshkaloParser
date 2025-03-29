import os
import re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

class AdParser:
    """Class to parse ads from a website and save them to an Excel file."""
    
    def __init__(self, config):
        """Initialize the AdParser with configuration settings."""
        self.options = Options()
        self.options.add_argument("--headless")
        self.service = Service(config['geckodriver_path'])
        self.driver = webdriver.Firefox(service=self.service, options=self.options)
        self.seen_links = set()
        self.min_price = config['min_price']
        self.max_price = config['max_price']
        self.max_square = config['max_square']
        self.save_dir = config['save_dir']
        os.makedirs(self.save_dir, exist_ok=True)

    def start_driver(self):
        """Start the web driver and remove ads from the page."""
        self.driver.get("https://www.njuskalo.hr")
        self.driver.execute_script("""
            var adSelectors = [
                'iframe[src*="amazon-adsystem.com"]',
                'iframe[src*="googlesyndication.com"]',
                'iframe[src*="googletagmanager.com"]',
                'iframe[src*="midas-network.com"]',
                'iframe[src*="privacy-center.org"]',
                'img[src*="defractal.com"]',
                'div[class*="ad"]',
                'script[src*="dotmetrics.net"]'
            ];
            adSelectors.forEach(function(selector) {
                var ads = document.querySelectorAll(selector);
                ads.forEach(function(ad) {
                    ad.remove();
                });
            });
        """)

    def get_element_text(self, ad, by, value):
        """Get the text of an element."""
        try:
            return ad.find_element(by, value).text.strip()
        except Exception as e:
            # print(f"Error getting text: {e}")
            return None

    def get_element_attr(self, ad, by, value, attr):
        """Get an attribute of an element."""
        try:
            return ad.find_element(by, value).get_attribute(attr)
        except Exception as e:
            # print(f"Error getting attribute: {e}")
            return None

    def wait_for_element(self, by, value, timeout=10):
        """Wait for an element to be present."""
        try:
            return WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((by, value))
            )
        except Exception as e:
            print(f"Error waiting for element: {e}")
            return None

    def parse_listings(self):
        """Parse the listings on the current page."""
        ads = self.driver.find_elements(By.CLASS_NAME, "EntityList-item")
        listings = []
        if not ads:
            return listings, 0
        for ad in ads:
            price = self.get_element_text(ad, By.CLASS_NAME, "ClassifiedDetailSummary-priceDomestic")
            if not price:
                price = self.get_element_text(ad, By.CLASS_NAME, "price")  # fallback to other class
            link = self.get_element_attr(ad, By.TAG_NAME, "a", "href")
            if not link or not link.startswith("https://www.njuskalo.hr/nekretnine/") or link in self.seen_links:
                continue
            self.seen_links.add(link)
            listings.append({"price": price, "link": link})
        return listings, len(listings)

    def collect_data(self, pages):
        """Collect data from multiple pages."""
        all_data = []
        total_ads = 0
        empty_pages = 0
        previous_links = self.load_previous_data()
        for page in range(1, pages):
            url = (f"https://www.njuskalo.hr/iznajmljivanje-stanova?"
                   f"geo[locationIds]=1248%2C1249%2C1250%2C1251%2C1252%2C1253&"
                   f"price[max]={self.max_price}&page={page}&"
                   f"livingArea[max]={self.max_square}")
            self.driver.get(url)
            data, _ = self.parse_listings()  # Ignore ad_count
            if not data:
                empty_pages += 1
                if empty_pages >= 2:
                    break
                continue
            unique_data = [ad for ad in data if ad["link"] not in previous_links]
            unique_count = len(unique_data)
            all_data.extend(unique_data)
            print(f"✅ Page {page}: {unique_count} unique listings found")
            total_ads += unique_count
            empty_pages = 0
        return all_data, total_ads

    def get_latest_file(self):
        """Get the latest Excel file from the save directory."""
        files = [f for f in os.listdir(self.save_dir) if f.endswith(".xlsx")]
        if not files:
            return None
        files.sort(key=lambda x: os.path.getmtime(os.path.join(self.save_dir, x)), reverse=True)
        return os.path.join(self.save_dir, files[0])

    def load_previous_data(self):
        """Load previously saved data from Excel files."""
        folder = self.save_dir
        files = sorted([f for f in os.listdir(folder) if f.endswith(".xlsx")],
                       key=lambda f: os.path.getmtime(os.path.join(folder, f)),
                       reverse=True)
        if not files:
            print("ℹ️ No previous data found.")
            return set()
        last_file = os.path.join(folder, files[0])
        try:
            wb = load_workbook(last_file, data_only=True)
            ws = wb.active
            previous_links = {row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[1]}
            wb.close()
            print(f"ℹ️ Loaded {len(previous_links)} previous listings from {last_file}")
            return previous_links
        except Exception as e:
            print(f"⚠️ Failed to load previous data: {e}")
            return set()

    def save_to_excel(self, data):
        """Save the collected data to an Excel file."""
        if not data:
            print("ℹ️ No data to save.")
            return
        date_str = datetime.now().strftime("%d %B %H-%M")
        folder = self.save_dir
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, f"njuskalo_listings {date_str}.xlsx")

        def extract_price(item):
            price = item["price"]
            if price is None:
                return float("inf")  # or some large number to handle missing prices
            match = re.search(r"\d+", price.replace(".", "").replace(",", ""))
            return int(match.group()) if match else float("inf")

        data.sort(key=extract_price)
        wb = Workbook()
        ws = wb.active
        ws.append(["Price", "Link"])
        for item in data:
            ws.append([item["price"], item["link"]])
        wb.save(file_path)
        print(f"✅ Data saved in {file_path}")

    def close_driver(self):
        """Close the web driver."""
        self.driver.quit()

if __name__ == "__main__":
    config = {
        'min_price': 250, 
        'max_price': 400, 
        'max_square': 45, 
        'geckodriver_path': "/usr/bin/geckodriver", 
        'save_dir': "data"
    }
    parser = AdParser(config)
    parser.start_driver()
    all_data, total_ads = parser.collect_data(100)
    print(f"✅ Total ads collected: {total_ads}")
    parser.save_to_excel(all_data)
    parser.close_driver()
